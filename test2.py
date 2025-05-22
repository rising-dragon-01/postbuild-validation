import os
from azure.identity import AzureCliCredential
from azure.mgmt.compute import ComputeManagementClient
from azure.mgmt.network import NetworkManagementClient
from azure.mgmt.resource import ResourceManagementClient
import openpyxl

# Config
subscription_id = os.popen("az account show --query id -o tsv").read().strip()
credential = AzureCliCredential()

compute_client = ComputeManagementClient(credential, subscription_id)
network_client = NetworkManagementClient(credential, subscription_id)
resource_client = ResourceManagementClient(credential, subscription_id)

def get_vm_basic_info(vm):
    return {
        "Location": vm.location,
        "Size": vm.hardware_profile.vm_size,
        "Operating System": vm.storage_profile.os_disk.os_type.value,
        "OS Disk Size (GB)": vm.storage_profile.os_disk.disk_size_gb,
        "Data Disk-1 Size (GB)": vm.storage_profile.data_disks[0].disk_size_gb if vm.storage_profile.data_disks else None,
        "Availability Set": vm.availability_set.id.split('/')[-1] if vm.availability_set else None,
        "Tags": str(vm.tags)
    }

def get_nic_and_ip_info(vm):
    nic_id = vm.network_profile.network_interfaces[0].id
    nic_rg = nic_id.split("/")[4]
    nic_name = nic_id.split("/")[-1]
    nic = network_client.network_interfaces.get(nic_rg, nic_name)
    private_ip = nic.ip_configurations[0].private_ip_address
    public_ip = None
    app_security_groups = [asg.id.split('/')[-1] for asg in nic.ip_configurations[0].application_security_groups] if nic.ip_configurations[0].application_security_groups else []
    subnet = nic.ip_configurations[0].subnet
    subnet_name = subnet.id.split('/')[-1]
    vnet_name = subnet.id.split('/')[8]
    vnet_rg = subnet.id.split('/')[4]
    if nic.ip_configurations[0].public_ip_address:
        public_ip_id = nic.ip_configurations[0].public_ip_address.id
        public_ip_name = public_ip_id.split('/')[-1]
        public_ip = network_client.public_ip_addresses.get(nic_rg, public_ip_name).ip_address
    return {
        "Assigned Private IP": private_ip,
        "Assigned Public IP": public_ip,
        "VNet Name": vnet_name,
        "VNet Resource Group": vnet_rg,
        "Subnet Name": subnet_name,
        "Application Security Groups": str(app_security_groups)
    }

def get_vm_extensions(resource_group, vm_name):
    extensions = compute_client.virtual_machine_extensions.list(resource_group, vm_name)
    extension_names = [ext.name for ext in extensions]
    return {"Extensions": str(extension_names)}

def get_backup_status(resource_group, vm):
    backup_status = os.popen(
        f"az backup item list --resource-group {resource_group} --vault-name <VAULT_NAME> --query \"[?properties.sourceResourceId=='{vm.id}'].properties.protectionState\" -o tsv"
    ).read().strip() or "Not Enabled"
    return {"Backup Status": backup_status}

def get_log_analytics(resource_group):
    log_analytics = os.popen(
        f"az monitor log-analytics workspace list --resource-group {resource_group} -o table"
    ).read().strip()
    return {"Log Analytics Workspaces": log_analytics}

def write_dict_to_excel(data, filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "VM Details"
    for idx, (key, value) in enumerate(data.items(), 1):
        ws.cell(row=idx, column=1, value=key)
        ws.cell(row=idx, column=2, value=value)
    wb.save(filename)

def get_vm_details(resource_group, vm_name):
    vm = compute_client.virtual_machines.get(resource_group, vm_name, expand='instanceView')
    details = {
        "VM": vm_name,
        "Resource Group": resource_group,
        "Subscription ID": subscription_id
    }
    details.update(get_vm_basic_info(vm))
    details.update(get_nic_and_ip_info(vm))
    details.update(get_vm_extensions(resource_group, vm_name))
    details.update(get_backup_status(resource_group, vm))
    details.update(get_log_analytics(resource_group))
    excel_filename = f"{vm_name}.xlsx"
    write_dict_to_excel(details, excel_filename)
    return excel_filename

# Example usage
# get_vm_details("MyResourceGroup", "MyVM")

if __name__ == "__main__":
    import sys
    if len(sys.argv) != 3:
        print("Usage: python test2.py <ResourceGroup> <VMName>")
        sys.exit(1)
    resource_group = sys.argv[1]
    vm_name = sys.argv[2]
    excel_file = get_vm_details(resource_group, vm_name)
    print(f"Excel file generated: {excel_file}")